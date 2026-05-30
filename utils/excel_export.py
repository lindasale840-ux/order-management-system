import io
import pandas as pd

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)

from openpyxl.utils import (
    get_column_letter
)


def dataframe_to_excel(
    dataframes: dict
):

    output = io.BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        for sheet_name, df in dataframes.items():

            df.to_excel(

                writer,

                sheet_name=sheet_name,

                index=False
            )

            worksheet = (
                writer.sheets[
                    sheet_name
                ]
            )

            # ======================
            # HEADER STYLE
            # ======================

            header_fill = PatternFill(

                fill_type="solid",

                fgColor="1F4E78"
            )

            header_font = Font(

                color="FFFFFF",

                bold=True
            )

            for cell in worksheet[1]:

                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(

                    horizontal="center"
                )

            # ======================
            # AUTO WIDTH
            # ======================

            for column in (
                worksheet.columns
            ):

                max_length = 0

                column_letter = (
                    get_column_letter(
                        column[0].column
                    )
                )

                for cell in column:

                    try:

                        max_length = max(

                            max_length,

                            len(
                                str(
                                    cell.value
                                )
                            )
                        )

                    except:

                        pass

                worksheet.column_dimensions[
                    column_letter
                ].width = (
                    max_length + 3
                )

            # ======================
            # FREEZE HEADER
            # ======================

            worksheet.freeze_panes = "A2"

    return output.getvalue()