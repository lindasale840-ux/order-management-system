import streamlit as st
import pandas as pd
from io import BytesIO
import time

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from utils.auth_guard import require_admin
from services.dashboard_service import DashboardService
from services.payment_service import PaymentService
from repositories.user_repository import UserRepository
from repositories.order_repository import OrderRepository

def generate_template():
    columns = [
        "customer_name", "order_number", "measurement_date", "cert_status", "sale_owner", "created_by",
        "disable_calibration_notification", "disable_document_notification", "disable_payment_notification", # Thêm cột mới
        "invoice_group", "invoice_date", "payment_terms", "payment_status", "total", "commission_percent", "note"
    ]

    df = pd.DataFrame(columns=columns)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Import", index=False)
        workbook = writer.book
        ws = writer.sheets["Import"]

        header_fill = PatternFill(fill_type="solid", start_color="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        required_fill = PatternFill(fill_type="solid", start_color="FFF2CC")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        required_cols = ["customer_name", "order_number", "sale_owner"]
        for col_name in required_cols:
            col_index = columns.index(col_name) + 1
            ws.cell(row=2, column=col_index).fill = required_fill

        sample_row = [
            "ABC Company", "GST001", "2026-06-01", "2026-06-05", "LINDA", "Thịnh",
            0, 0, 0, "INV001", "2026-06-10", 30, "2026-07-05", 5000000, 10, "Historical import example"
        ]

        for col_num, value in enumerate(sample_row, start=1):
            ws.cell(row=2, column=col_num).value = value

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = min(length + 5, 40)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        payment_terms_validation = DataValidation(type="list", formula1='"30,45,60,90"')
        ws.add_data_validation(payment_terms_validation)
        payment_terms_validation.add("L2:L5000") # Dịch chuyển cột validation từ K sang L do chèn thêm cột

        instruction_ws = workbook.create_sheet("Instruction")
        instruction_ws.append(["Field", "Required", "Description", "Example"])

        instructions = [
            ["customer_name","YES","Customer Name","ABC Company"],
            ["order_number","YES","Order Number","GST001"],
            ["measurement_date","NO","Calibration Date","2026-06-01"],
            ["cert_status","NO","Certificate Date","2026-06-05"],
            ["sale_owner","YES","Sales Owner","LINDA"],
            ["created_by","YES","Assistant Owner","Thịnh"],
            ["disable_calibration_notification","NO","0=Track, 1=Disable notification","0"],
            ["disable_document_notification","NO","0=Track, 1=Disable document notification","0"],
            ["disable_payment_notification","NO","0=Track, 1=Disable payment overdue notification","0"], # Chỉ dẫn cột mới
            ["invoice_group","NO","Invoice Group","INV001"],
            ["invoice_date","NO","Invoice Date","2026-06-10"],
            ["payment_terms","NO","Payment Term","30"],
            ["payment_status","NO","Paid Date","2026-07-05"],
            ["total","NO","Revenue","5000000"],
            ["commission_percent","NO","Commission %","10"],
            ["note","NO","Remark","Historical import"]
        ]

        for row in instructions:
            instruction_ws.append(row)

        for cell in instruction_ws[1]:
            cell.fill = header_fill
            cell.font = header_font

    return output.getvalue()


def show_historical_import_page():
    require_admin()

    st.title("📥 Historical Data Import")

    template_data = generate_template()

    st.download_button(
        label="📥 Download Import Template",
        data=template_data,
        file_name="ERP_Import_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    uploaded_file = st.file_uploader("Upload Historical Excel", type=["xlsx"])

    if uploaded_file is None:
        if "historical_df" in st.session_state:
            del st.session_state["historical_df"]
        st.info("Please upload Excel file.")
        return

    if "historical_df" not in st.session_state:
        st.session_state["historical_df"] = pd.read_excel(uploaded_file)
        st.success(f"{len(st.session_state['historical_df'])} rows loaded into memory!")

    df = st.session_state["historical_df"]
    
    users_df = UserRepository.get_all_users()
    valid_users = users_df["username"].tolist()
    
    preview_rows = []
    for _, row in df.iterrows():
        order_number = row.get("order_number")
        existing = OrderRepository.get_by_order_number(order_number)
        action = "UPDATE" if not existing.empty else "NEW"

        preview_rows.append({
            "order_number": order_number,
            "customer_name": row.get("customer_name"),
            "sale_owner": row.get("sale_owner"),
            "assistant": row.get("created_by"),
            "action": action
        })

    preview_df = pd.DataFrame(preview_rows)
    
    new_count = len(preview_df[preview_df["action"] == "NEW"])
    update_count = len(preview_df[preview_df["action"] == "UPDATE"])
    total_revenue = pd.to_numeric(df["total"], errors="coerce").fillna(0).sum()

    col1, col2, col3 = st.columns(3)
    col1.metric("NEW Orders", new_count)
    col2.metric("UPDATE Orders", update_count)
    col3.metric("Revenue Import", f"{total_revenue:,.0f}")

    st.subheader("📋 Import Preview")
    st.dataframe(preview_df, use_container_width=True)
    
    if "import_processing" not in st.session_state:
        st.session_state["import_processing"] = False

    with st.form(key="import_form_v2"):
        st.markdown("### 🛠️ Import Actions")
        validate_clicked = st.form_submit_button("🔍 Validate File")
        confirm_import = st.checkbox("I confirm importing historical data")
        import_clicked = st.form_submit_button("🚀 Import To ERP")

    if validate_clicked:
        validation_errors = []
        if df["customer_name"].isna().any(): validation_errors.append("Missing customer_name")
        if df["order_number"].isna().any(): validation_errors.append("Missing order_number")
        if df["sale_owner"].isna().any(): validation_errors.append("Missing sale_owner")

        duplicates = df[df["order_number"].duplicated(keep=False)]
        if not duplicates.empty:
            validation_errors.append(f"Duplicate order_number found: {duplicates['order_number'].tolist()}")

        invalid_created_by_rows = df[~df["created_by"].astype(str).isin(valid_users)]
        if not invalid_created_by_rows.empty:
            wrong_names = invalid_created_by_rows["created_by"].unique().tolist()
            validation_errors.append(
                f"Invalid username found: {wrong_names}. Hệ thống yêu cầu điền chính xác 'username' hệ thống."
            )
            st.info(f"💡 Danh sách Username hợp lệ đang có trên ERP của bạn: `{valid_users}`")

        if validation_errors:
            st.error("Validation Failed")
            for err in validation_errors:
                st.write("❌", err)
        else:
            st.success("Validation Passed! Data hợp lệ 100%.")

    if import_clicked:
        if not confirm_import:
            st.error("Please check 'I confirm...' inside the form first!")
            return
        st.session_state["import_processing"] = True

    if st.session_state["import_processing"]:
        imported_orders = 0
        imported_payments = 0
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        total_rows = len(df)

        if "processed_indices" not in st.session_state:
            st.session_state["processed_indices"] = set()

        for index, row in df.iterrows():
            if index in st.session_state["processed_indices"]:
                imported_orders += 1
                imported_payments += 1
                continue

            if index % 5 == 0 or (index + 1) == total_rows:
                percent_complete = int((index + 1) / total_rows * 100)
                progress_bar.progress(percent_complete)
                status_text.text(f"🚀 Đang nạp dữ liệu: Dòng {index + 1}/{total_rows}...")

            measurement_date = pd.to_datetime(row.get("measurement_date")).date().isoformat() if pd.notna(row.get("measurement_date")) else None
            cert_status = pd.to_datetime(row.get("cert_status")).date().isoformat() if pd.notna(row.get("cert_status")) else None
            invoice_date = pd.to_datetime(row.get("invoice_date")).date().isoformat() if pd.notna(row.get("invoice_date")) else None
            
            raw_payment_status = row.get("payment_status")
            if pd.isna(raw_payment_status) or str(raw_payment_status).strip() == "" or str(raw_payment_status).lower() in ["nan", "none"]:
                payment_status = None
            else:
                try:
                    payment_status = pd.to_datetime(raw_payment_status).date().isoformat()
                except Exception:
                    payment_status = str(raw_payment_status).strip()
            
            created_by = row.get("created_by") if pd.notna(row.get("created_by")) else st.session_state["username"]
            disable_notification = int(row.get("disable_calibration_notification", 0) or 0)
            disable_document_notification = int(row.get("disable_document_notification", 0) or 0)
            disable_payment_notification = int(row.get("disable_payment_notification", 0) or 0) # Lấy dữ liệu cấu hình mới từ dòng Excel

            try:
                # 1. Đồng bộ hóa thông tin đơn hàng cùng tham số cấu hình thông báo mới
                DashboardService.sync_order(
                    customer_name=row.get("customer_name"),
                    order_number=row.get("order_number"),
                    measurement_date=measurement_date,
                    cert_status=cert_status,
                    sale_owner=row.get("sale_owner"),
                    created_by=created_by,
                    disable_calibration_notification=disable_notification,
                    disable_document_notification=disable_document_notification,
                    disable_payment_notification=disable_payment_notification
                )

                # 2. Đồng bộ hóa thông tin hóa đơn
                total = float(row.get("total", 0) or 0)
                commission_percent = float(row.get("commission_percent", 0) or 0)
                payment_terms = row.get("payment_terms", 0)
                if pd.isna(payment_terms):
                    payment_terms = 0
                            
                PaymentService.save_invoice(
                    order_number=row.get("order_number"),
                    invoice_date=invoice_date,
                    invoice_group=row.get("invoice_group"),
                    payment_terms=payment_terms,
                    payment_status=payment_status,
                    total=total,
                    commission_percent=commission_percent,
                    note=row.get("note"),
                    invoice_created_by=created_by
                )
                
                st.session_state["processed_indices"].add(index)
                imported_orders += 1
                imported_payments += 1
                time.sleep(0.01)

            except Exception as e:
                print(f"Error at row {index + 2}: {str(e)}")
                continue

        st.cache_data.clear()
        status_text.empty()
        progress_bar.empty()
        
        if "historical_df" in st.session_state:
            del st.session_state["historical_df"]
        if "processed_indices" in st.session_state:
            del st.session_state["processed_indices"]
            
        st.session_state["import_processing"] = False
        st.success(f"🎉 Hệ thống đã đồng bộ thành công hoàn toàn {imported_orders}/{total_rows} dòng dữ liệu lịch sử.")
        time.sleep(1.5)
        st.rerun()